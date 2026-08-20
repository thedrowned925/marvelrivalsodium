#!/usr/bin/env python3
from __future__ import annotations
import argparse,datetime as dt,hashlib,json,re,tempfile,unicodedata
from pathlib import Path
from typing import Any
import openpyxl,requests

STATUS={"Bekliyor":"waiting","Kayit Alindi":"recorded","Kayıt Alındı":"recorded","Kontrol Edildi":"checked","Oyuna Eklendi":"added"}
COLUMNS=["Sıra","WAV Dosya Adı","WEM ID","Internal Name","English","Türkçe","Durum","Seslendiren","Tarih","Not"]

def download(url:str,target:Path):
    with requests.get(url,stream=True,timeout=60,allow_redirects=True) as r:
        r.raise_for_status()
        with target.open('wb') as f:
            for chunk in r.iter_content(262144):
                if chunk:f.write(chunk)
    if target.stat().st_size<50000:raise RuntimeError('Downloaded workbook is unexpectedly small')

def slug(v:str)->str:
    v=unicodedata.normalize('NFKD',v).encode('ascii','ignore').decode().lower()
    return re.sub(r'[^a-z0-9]+','-',v).strip('-') or 'character'

def j(v:Any):
    if v is None:return None
    if isinstance(v,(dt.datetime,dt.date,dt.time)):return v.isoformat()
    if isinstance(v,float) and v.is_integer():return int(v)
    return v

def rows_for(wb,name:str):
    if name not in wb.sheetnames:return []
    out=[]
    for a in wb[name].iter_rows(min_row=2,min_col=1,max_col=10,values_only=True):
        if not any(v is not None and str(v).strip() for v in a[:6]):continue
        v=[j(x) for x in a]
        out.append({"order":v[0],"wav":v[1],"wemId":v[2],"internalName":v[3],"english":v[4],"turkish":v[5],"status":v[6],"voiceActor":v[7],"date":v[8],"note":v[9]})
    return out

def counts(rows,fallback):
    c={"waiting":0,"recorded":0,"checked":0,"added":0}
    for r in rows:
        k=STATUS.get(str(r.get('status')).strip()) if r.get('status') is not None else None
        if k:c[k]+=1
    return fallback if sum(c.values())==0 and sum(fallback.values()) else c

def build(book:Path,details:Path,existing=None):
    wb=openpyxl.load_workbook(book,data_only=True,read_only=True)
    if 'Karakter Bazli' not in wb.sheetnames:raise RuntimeError("'Karakter Bazli' sheet not found")
    details.mkdir(parents=True,exist_ok=True);chars=[];payloads=[];revision=[];expected=set()
    for r in wb['Karakter Bazli'].iter_rows(min_row=2,max_col=9,values_only=True):
        if not r[0]:continue
        name=str(r[0]);line_rows=rows_for(wb,name);total=len(line_rows) or int(r[1] or 0)
        fb={"waiting":int(r[2] or 0),"recorded":int(r[3] or 0),"checked":int(r[4] or 0),"added":int(r[5] or 0)};c=counts(line_rows,fb)
        worked=c['recorded']+c['checked']+c['added'];progress=round(worked/total*100,2) if total else 0
        status='Tamamlandi' if total and c['added']==total else ('Baslamadi' if total and c['waiting']==total else 'Devam Ediyor')
        actor=r[8] if r[8] else None;s=slug(name);path=f'data/characters/{s}.json'
        char={"name":name,"slug":s,"detailPath":path,"total":total,**c,"worked":worked,"progress":progress,"status":status,"voiceActor":actor};chars.append(char)
        detail={"character":name,"voiceActor":actor,"summary":{"total":total,**c,"worked":worked,"progress":progress,"status":status},"columns":COLUMNS,"rows":line_rows};target=details/f'{s}.json';expected.add(target);payloads.append((target,detail));revision.append({"character":char,"rows":line_rows})
    rev=hashlib.sha256(json.dumps(revision,ensure_ascii=False,sort_keys=True,separators=(',',':')).encode()).hexdigest()[:16]
    generated=(existing or {}).get('generatedAt') if (existing or {}).get('dataRevision')==rev else dt.datetime.now(dt.timezone.utc).isoformat().replace('+00:00','Z')
    for target,d in payloads:
        d['dataRevision']=rev;d['generatedAt']=generated;text=json.dumps(d,ensure_ascii=False,indent=2)+'\n'
        if not target.exists() or target.read_text('utf-8')!=text:target.write_text(text,encoding='utf-8')
    for old in details.glob('*.json'):
        if old not in expected:old.unlink()
    total=sum(x['total'] for x in chars);stats={k:sum(x[k] for x in chars) for k in ('waiting','recorded','checked','added')};worked=stats['recorded']+stats['checked']+stats['added'];sc={}
    for x in chars:sc[x['status']]=sc.get(x['status'],0)+1
    source=None
    if 'Genel' in wb.sheetnames and wb['Genel']['B4'].value:source=str(wb['Genel']['B4'].value)
    return {"project":"Marvel Rivals Türkçe Dublaj Projesi","managers":["Efe Karatepe","Hasan Çağın Yıldırım"],"sourceWorkbookUpdatedText":source,"dataRevision":rev,"generatedAt":generated,"stats":{"characters":len(chars),"totalLines":total,**stats,"worked":worked,"progress":round(worked/total*100,2) if total else 0,"statusCounts":sc,"assigned":sum(1 for x in chars if x['voiceActor'])},"characters":chars}

def main():
    p=argparse.ArgumentParser();p.add_argument('--url',required=True);p.add_argument('--output',default='data/data.json');p.add_argument('--details-dir',default='data/characters');a=p.parse_args();out=Path(a.output);out.parent.mkdir(parents=True,exist_ok=True);existing=None
    if out.exists():
        try:existing=json.loads(out.read_text('utf-8'))
        except Exception:pass
    with tempfile.TemporaryDirectory() as td:
        x=Path(td)/'source.xlsx';download(a.url,x);data=build(x,Path(a.details_dir),existing)
    out.write_text(json.dumps(data,ensure_ascii=False,indent=2)+'\n',encoding='utf-8');print(f"{data['stats']['characters']} characters / {data['stats']['totalLines']} lines / revision {data['dataRevision']}")
if __name__=='__main__':main()
